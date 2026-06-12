import os
import datetime
import json

from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session
from werkzeug.utils import secure_filename
import pandas as pd
import numpy as np

from backend.data_loader import load_data
from backend.data_cleaning import clean_dataset, analyze_quality, get_cleaning_summary
from backend.preprocessing import detect_data_types
from backend.descriptive_stats import get_summary_metrics, get_descriptive_stats
from backend.advanced_stats import get_advanced_stats
from backend.dashboard_overview import generate_overview_dashboard
from backend.viz_engine import (
    generate_master_chart, category_available, CATEGORY_CHARTS, CHART_LABELS,
)
from backend.insights_generator import generate_auto_insights
from backend.time_series import detect_datetime_cols, generate_ts_plots
from backend.cleaning_engine import (
    get_session, reset_session, delete_session,
    get_quality_report, detect_data_status,
)
from backend.data_sanitizer import sanitize_df_numeric_cols

app = Flask(__name__, template_folder='frontend/templates', static_folder='frontend/static')
app.secret_key = 'super_secret_key_week_15_itsb'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024

ALLOWED_EXTENSIONS = {'csv', 'txt', 'xlsx'}
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def _get_or_init_session(filename):
    """
    Ambil CleaningSession untuk filename.
    Jika belum ada, load dari disk dan inisialisasi.
    Secures against path traversal attacks by validating with secure_filename.
    Return: (session, error_msg)
    """
    safe_name = secure_filename(filename)
    if not safe_name or safe_name != filename:
        return None, 'Nama file tidak valid.'

    sess = get_session(safe_name)
    if sess is not None:
        return sess, None

    base_dir = os.path.abspath(app.config['UPLOAD_FOLDER'])
    filepath = os.path.abspath(os.path.join(base_dir, safe_name))
    if not filepath.startswith(base_dir):
        return None, 'Akses tidak sah.'

    if not os.path.exists(filepath):
        return None, 'File tidak ditemukan.'

    df_raw = load_data(filepath)
    if df_raw is None:
        return None, 'Gagal membaca dataset.'

    sess = reset_session(safe_name, df_raw)
    return sess, None


# ─────────────────────────────────────────────────────────────────────────────
# ROUTES
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/upload', methods=['GET', 'POST'])
def upload_file():
    uploaded_files = []
    try:
        if os.path.exists(app.config['UPLOAD_FOLDER']):
            folder_path = app.config['UPLOAD_FOLDER']
            files = [
                f for f in os.listdir(folder_path)
                if os.path.isfile(os.path.join(folder_path, f))
            ]
            files.sort(
                key=lambda x: os.path.getmtime(os.path.join(folder_path, x)),
                reverse=True,
            )
            uploaded_files = files
    except Exception as e:
        print(f"Error membaca history: {e}")

    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file part', 'error')
            return redirect(request.url)

        file = request.files['file']
        if file.filename == '':
            flash('No selected file', 'error')
            return redirect(request.url)

        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

            # ── Deteksi duplikat ──────────────────────────────────────────────
            if os.path.exists(filepath):
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return jsonify({
                        'duplicate': True,
                        'filename' : filename,
                        'redirect' : url_for('dashboard', filename=filename),
                    })
                flash(
                    f'Dataset "{filename}" sudah pernah dianalisis sebelumnya. '
                    'Menampilkan data lama...', 'info',
                )
                return redirect(url_for('dashboard', filename=filename))

            file.save(filepath)

            # ── Init cleaning session untuk file baru ─────────────────────────
            df_raw = load_data(filepath)
            if df_raw is not None:
                reset_session(filename, df_raw)

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({
                    'duplicate': False,
                    'filename' : filename,
                    'redirect' : url_for('dashboard', filename=filename),
                })

            flash(f'Dataset "{filename}" berhasil diunggah!', 'success')
            return redirect(url_for('dashboard', filename=filename))
        else:
            flash('Format file tidak valid. Gunakan .csv, .txt, atau .xlsx', 'error')

    return render_template('upload.html', history=uploaded_files)


def _get_dashboard_context(filename):
    """
    Computes dashboard context variables.
    Secures against path traversal attacks.
    """
    safe_name = secure_filename(filename)
    if not safe_name or safe_name != filename:
        return None

    base_dir = os.path.abspath(app.config['UPLOAD_FOLDER'])
    filepath = os.path.abspath(os.path.join(base_dir, safe_name))
    if not filepath.startswith(base_dir):
        return None

    if not os.path.exists(filepath):
        return None

    df_raw = load_data(filepath)
    if df_raw is None:
        return None

    # ── Ambil atau inisialisasi CleaningSession ───────────────────────────────
    sess = get_session(safe_name)
    if sess is None:
        sess = reset_session(safe_name, df_raw)

    # ── Tentukan df yang dipakai (cleaned atau raw) ───────────────────────────
    is_cleaned   = sess.is_cleaned
    df           = sess.df_current     # cleaned kalau sudah cleaning, raw kalau belum
    data_status  = detect_data_status(df)

    # ── Quality report (dari current df) untuk overview warning ─────────────────────
    quality_full = get_quality_report(df)

    # ── Backward-compat: cleaning summary & log ───────────────────────────────
    if is_cleaned and len(sess._history) > 1:
        last_snap      = sess._history[-1]
        cleaning_log   = last_snap.summary.get('log', [])
        cleaning_summary = {
            'rows_before'       : last_snap.summary['rows_before'],
            'rows_after'        : last_snap.summary['rows_after'],
            'cols_before'       : last_snap.summary['cols_before'],
            'cols_after'        : last_snap.summary['cols_after'],
            'missing_before'    : last_snap.summary['missing_before'],
            'missing_after'     : last_snap.summary['missing_after'],
            'missing_pct_before': round(
                last_snap.summary['missing_before'] /
                max(last_snap.summary['rows_before'] * last_snap.summary['cols_before'], 1) * 100, 2
            ),
            'missing_pct_after' : round(
                last_snap.summary['missing_after'] /
                max(last_snap.summary['rows_after'] * last_snap.summary['cols_after'], 1) * 100, 2
            ),
            'duplicates_removed': last_snap.summary.get('rows_removed', 0),
            'log'               : cleaning_log,
        }
    else:
        df_temp, cleaning_log_temp = clean_dataset(df_raw, {})
        cleaning_summary = get_cleaning_summary(df_raw, df_temp, cleaning_log_temp)

    quality_report = analyze_quality(df)

    # ── Column detection ──────────────────────────────────────────────────────
    num_cols, cat_cols = detect_data_types(df)

    # ── GLOBAL DATA SANITIZATION ──────────────────────────────────────────────
    df_stats = sanitize_df_numeric_cols(df, num_cols)

    # ── Summary metrics ───────────────────────────────────────────────────────
    metrics = get_summary_metrics(df_stats, num_cols, cat_cols)

    # ── Descriptive stats ─────────────────────────────────────────────────────
    num_stats, cat_stats = get_descriptive_stats(df_stats, num_cols, cat_cols)

    # ── Advanced stats ────────────────────────────────────────────────────────
    advanced = get_advanced_stats(df_stats, num_cols, cat_cols)

    # ── Time Series detection ─────────────────────────────────────────────────
    dt_cols = detect_datetime_cols(df)

    # ── Dashboard overview ────────────────────────────────────────────────────
    overview = generate_overview_dashboard(
        df, num_cols, cat_cols, dt_cols=dt_cols, metrics=metrics
    )

    # ── Time Series plots ─────────────────────────────────────────────────────
    plots            = {}
    ts_insights_list = []
    ts_meta          = {}

    if dt_cols:
        try:
            ts_plots, ts_insights_list, ts_meta = generate_ts_plots(
                df, dt_cols, num_cols
            )
            plots = ts_plots
        except Exception as e:
            print(f"[TS] generate_ts_plots error: {e}")

    # ── Intelligent insights (bilingual) ───────────────────────────────────
    auto_insights_bilingual = generate_auto_insights(
        df, num_cols, cat_cols, ts_insights=ts_insights_list
    )
    auto_insights_en = auto_insights_bilingual.get('en', [])
    auto_insights_id = auto_insights_bilingual.get('id', [])

    # ── Preview HTML ──────────────────────────────────────────────────────────
    preview_raw_html = df_raw.to_html(
        classes='data-table display nowrap',
        index=False,
        table_id='preview-raw-table',
    )
    preview_clean_html = df.to_html(
        classes='data-table display nowrap',
        index=False,
        table_id='preview-clean-table',
    )

    # ── Dataset info untuk sidebar card ───────────────────────────────────────
    file_size_kb = 0.0
    if os.path.exists(filepath):
        file_size_kb = os.path.getsize(filepath) / 1024.0
    file_size_str = f"{file_size_kb:,.2f} KB"

    dt_file = datetime.datetime.fromtimestamp(os.path.getmtime(filepath))
    months  = ["Jan","Feb","Mar","Apr","May","Jun",
               "Jul","Aug","Sep","Oct","Nov","Dec"]
    upload_time_str = (
        f"{dt_file.day:02d} {months[dt_file.month - 1]} {dt_file.year} "
        f"{dt_file.hour:02d}:{dt_file.minute:02d}"
    )
    dataset_info = {
        'name': safe_name,
        'rows': metrics['total_rows'],
        'cols': metrics['total_columns'],
        'size': file_size_str,
        'time': upload_time_str,
    }

    # ── col_data untuk interactive chart (sampled) ────────────────────────────
    col_data = {}
    for col in num_cols:
        series = df_stats[col].dropna()
        if len(series) > 2000:
            series = series.sample(2000, random_state=42)
        
        vals = []
        for v in series.tolist():
            try:
                f = float(v)
                if not np.isnan(f) and not np.isinf(f):
                    vals.append(round(f, 4))
            except (ValueError, TypeError):
                continue
        col_data[col] = {
            'type'  : 'numeric',
            'values': vals,
        }
    for col in cat_cols:
        vc = df[col].value_counts().head(20)
        col_data[col] = {
            'type'  : 'categorical',
            'labels': vc.index.astype(str).tolist(),
            'counts': [int(v) for v in vc.values.tolist()],
        }

    # ── Cleaning history untuk UI ─────────────────────────────────────────────
    cleaning_history = sess.history_labels

    # ── clean_opts dummy untuk backward-compat template lama ─────────────────
    clean_opts = {
        'strip_whitespace'       : True,
        'empty_to_nan'           : True,
        'drop_duplicates'        : True,
        'drop_empty_cols'        : True,
        'drop_high_missing'      : 0.0,
        'fill_missing_numeric'   : 'none',
        'fill_missing_categorical': 'none',
        'cap_outliers'           : False,
    }

    return {
        'filename'          : safe_name,
        'data_status'       : data_status,
        'preview'           : preview_clean_html if is_cleaned else preview_raw_html,
        'preview_raw'       : preview_raw_html,
        'preview_clean'     : preview_clean_html,
        'is_cleaned'        : is_cleaned,
        'metrics'           : metrics,
        'num_stats'         : num_stats,
        'cat_stats'         : cat_stats,
        'plots'             : plots,
        'insights'          : auto_insights_en,
        'insights_en'       : auto_insights_en,
        'insights_id'       : auto_insights_id,
        'advanced'          : advanced,
        'num_cols'          : num_cols,
        'cat_cols'          : cat_cols,
        'col_data'          : col_data,
        'dataset_info'      : dataset_info,
        'cleaning_summary'  : cleaning_summary,
        'quality_report'    : quality_report,
        'quality_full'      : quality_full,
        'clean_opts'        : clean_opts,
        'cleaning_history'  : cleaning_history,
        'overview'          : overview,
        'has_ts'            : bool(dt_cols),
        'dt_cols'           : dt_cols,
        'ts_meta'           : ts_meta,
    }


@app.route('/dashboard/<filename>')
def dashboard(filename):
    safe_name = secure_filename(filename)
    if not safe_name or safe_name != filename:
        flash('Nama file tidak valid.', 'error')
        return redirect(url_for('upload_file'))

    ctx = _get_dashboard_context(safe_name)
    if ctx is None:
        flash('File atau data tidak dapat dimuat.', 'error')
        return redirect(url_for('upload_file'))

    return render_template('dashboard.html', **ctx)


@app.route('/dashboard/<filename>/report/pdf')
def download_report_pdf(filename):
    safe_name = secure_filename(filename)
    if not safe_name or safe_name != filename:
        flash('Nama file tidak valid.', 'error')
        return redirect(url_for('upload_file'))

    ctx = _get_dashboard_context(safe_name)
    if ctx is None:
        flash('File atau data tidak dapat dimuat.', 'error')
        return redirect(url_for('upload_file'))

    # Path to temp PDF file
    temp_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'temp_reports')
    os.makedirs(temp_dir, exist_ok=True)
    temp_pdf_path = os.path.join(temp_dir, f"{safe_name}_report.pdf")

    # Load sess to pass df
    sess = get_session(safe_name)
    df = sess.df_current

    from backend.report_generator import generate_pdf_report
    try:
        generate_pdf_report(
            temp_pdf_path,
            safe_name,
            df,
            ctx['quality_report'],
            ctx['metrics'],
            ctx['num_stats'],
            ctx['cat_stats'],
            ctx['insights'],
            ctx['cleaning_history'],
            ctx['cleaning_summary']
        )
        from flask import send_file
        return send_file(
            temp_pdf_path,
            as_attachment=True,
            download_name=f"{os.path.splitext(safe_name)[0]}_report.pdf"
        )
    except Exception as e:
        flash(f"Gagal membuat laporan PDF: {e}", 'error')
        return redirect(url_for('dashboard', filename=safe_name))


def inline_html_assets(rendered_html):
    """
    Inlines style.css, jquery.min.js, plotly.min.js, script.js, dashboardOverview.js,
    and visualizationsMaster.js into the rendered HTML to ensure fully functional offline support.
    """
    static_dir = os.path.join(app.root_path, 'frontend', 'static')
    import re

    # 1. Inline style.css
    css_path = os.path.join(static_dir, 'css', 'style.css')
    if os.path.exists(css_path):
        try:
            with open(css_path, 'r', encoding='utf-8') as f:
                css_content = f.read()
            css_tag = f'<style>\n/* INLINED style.css */\n{css_content}\n</style>'
            rendered_html = re.sub(
                r'<link[^>]*href=[^>]*style\.css[^>]*>',
                css_tag,
                rendered_html
            )
        except Exception as e:
            print(f"[HTML Export] Error inlining CSS: {e}")

    # 2. Inline Javascript files
    js_files = [
        (r'jquery\.min\.js', os.path.join(static_dir, 'js', 'jquery.min.js')),
        (r'plotly\.min\.js', os.path.join(static_dir, 'js', 'plotly.min.js')),
        (r'script\.js', os.path.join(static_dir, 'js', 'script.js')),
        (r'dashboardOverview\.js', os.path.join(static_dir, 'js', 'dashboardOverview.js')),
        (r'visualizationsMaster\.js', os.path.join(static_dir, 'js', 'visualizationsMaster.js')),
    ]

    for pattern, file_path in js_files:
        if os.path.exists(file_path):
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    js_content = f.read()
                # Use a lambda replacement to prevent backslash issues in file contents during re.sub
                # Escape backslashes for JS strings safely
                js_tag = f'<script type="text/javascript">\n/* INLINED {os.path.basename(file_path)} */\n{js_content}\n</script>'
                
                # We do simple string replacement if possible to avoid regex overhead on massive files (like Plotly)
                # Let's find script tags matching pattern
                tag_pattern = re.compile(r'<script[^>]*src=[^>]*' + pattern + r'[^>]*>\s*</script>')
                rendered_html = tag_pattern.sub(lambda m: js_tag, rendered_html)
            except Exception as e:
                print(f"[HTML Export] Error inlining {os.path.basename(file_path)}: {e}")

    return rendered_html


@app.route('/dashboard/<filename>/report/html')
def download_report_html(filename):
    safe_name = secure_filename(filename)
    if not safe_name or safe_name != filename:
        flash('Nama file tidak valid.', 'error')
        return redirect(url_for('upload_file'))

    ctx = _get_dashboard_context(safe_name)
    if ctx is None:
        flash('File atau data tidak dapat dimuat.', 'error')
        return redirect(url_for('upload_file'))

    rendered = render_template('dashboard.html', **ctx)
    inlined = inline_html_assets(rendered)

    from flask import Response
    return Response(
        inlined,
        mimetype='text/html',
        headers={"Content-disposition": f"attachment; filename={os.path.splitext(safe_name)[0]}_dashboard.html"}
    )


@app.route('/dashboard/<filename>/export/excel')
def export_excel(filename):
    safe_name = secure_filename(filename)
    if not safe_name or safe_name != filename:
        flash('Nama file tidak valid.', 'error')
        return redirect(url_for('upload_file'))

    sess, err = _get_or_init_session(safe_name)
    if err:
        flash(err, 'error')
        return redirect(url_for('upload_file'))

    df = sess.df_current

    from io import BytesIO
    from openpyxl.utils import get_column_letter

    excel_buffer = BytesIO()
    try:
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Cleaned Data', index=False)
            workbook = writer.book
            worksheet = writer.sheets['Cleaned Data']

            # Enable AutoFilter on header row
            max_col = worksheet.max_column
            max_row = worksheet.max_row
            if max_col > 0 and max_row > 0:
                last_col_letter = get_column_letter(max_col)
                worksheet.auto_filter.ref = f"A1:{last_col_letter}{max_row}"

            # Auto-adjust column widths
            for col in worksheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                worksheet.column_dimensions[col_letter].width = max(max_len + 3, 10)

        excel_buffer.seek(0)
        from flask import send_file
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"{os.path.splitext(safe_name)[0]}_cleaned.xlsx"
        )
    except Exception as e:
        flash(f"Gagal mengekspor data ke Excel: {e}", 'error')
        return redirect(url_for('dashboard', filename=safe_name))


@app.route('/dashboard/<filename>/export/csv')
def export_csv(filename):
    safe_name = secure_filename(filename)
    if not safe_name or safe_name != filename:
        flash('Nama file tidak valid.', 'error')
        return redirect(url_for('upload_file'))

    sess, err = _get_or_init_session(safe_name)
    if err:
        flash(err, 'error')
        return redirect(url_for('upload_file'))

    df = sess.df_current

    from io import BytesIO, StringIO
    try:
        str_io = StringIO()
        df.to_csv(str_io, index=False)
        mem = BytesIO(str_io.getvalue().encode('utf-8'))
        from flask import send_file
        return send_file(
            mem,
            mimetype='text/csv',
            as_attachment=True,
            download_name=f"{os.path.splitext(safe_name)[0]}_cleaned.csv"
        )
    except Exception as e:
        flash(f"Gagal mengekspor data ke CSV: {e}", 'error')
        return redirect(url_for('dashboard', filename=safe_name))


# ─────────────────────────────────────────────────────────────────────────────
# CLEANING API ROUTES
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/api/cleaning/status/<filename>')
def api_cleaning_status(filename):
    """
    GET status session cleaning:
    - is_cleaned, quality_report dari df_current
    - history labels
    - rows & cols current
    """
    sess, err = _get_or_init_session(filename)
    if err:
        return jsonify({'ok': False, 'error': err}), 404

    df = sess.df_current
    return jsonify({
        'ok'          : True,
        'is_cleaned'  : sess.is_cleaned,
        'rows'        : len(df),
        'cols'        : len(df.columns),
        'history'     : sess.history_labels,
        'quality'     : get_quality_report(df),
        'missing_total': int(df.isna().sum().sum()),
    })


@app.route('/api/cleaning/preview/<filename>', methods=['POST'])
def api_cleaning_preview(filename):
    """
    POST preview dampak 1 operasi cleaning TANPA apply.
    Body JSON: { op_type, op_params }
    """
    sess, err = _get_or_init_session(filename)
    if err:
        return jsonify({'ok': False, 'error': err}), 404

    data     = request.get_json() or {}
    op_type  = data.get('op_type', '')
    op_params = data.get('op_params', {})

    if not op_type:
        return jsonify({'ok': False, 'error': 'op_type diperlukan'}), 400

    result = sess.preview_step(op_type, op_params)
    return jsonify(result)


@app.route('/api/cleaning/apply/<filename>', methods=['POST'])
def api_cleaning_apply(filename):
    """
    POST apply 1 operasi cleaning dan simpan ke history.
    Body JSON: { op_type, op_params }
    """
    sess, err = _get_or_init_session(filename)
    if err:
        return jsonify({'ok': False, 'error': err}), 404

    data      = request.get_json() or {}
    op_type   = data.get('op_type', '')
    op_params = data.get('op_params', {})

    if not op_type:
        return jsonify({'ok': False, 'error': 'op_type diperlukan'}), 400

    result = sess.apply_step(op_type, op_params)
    return jsonify(result)


@app.route('/api/cleaning/undo/<filename>', methods=['POST'])
def api_cleaning_undo(filename):
    """
    POST undo 1 langkah terakhir.
    Body JSON opsional: { index } untuk undo ke titik tertentu
    """
    sess, err = _get_or_init_session(filename)
    if err:
        return jsonify({'ok': False, 'error': err}), 404

    data  = request.get_json() or {}
    index = data.get('index')

    if index is not None:
        result = sess.undo_to(int(index))
    else:
        result = sess.undo()

    return jsonify(result)


@app.route('/api/cleaning/reset/<filename>', methods=['POST'])
def api_cleaning_reset(filename):
    """POST reset ke raw data."""
    sess, err = _get_or_init_session(filename)
    if err:
        return jsonify({'ok': False, 'error': err}), 404

    result = sess.reset()
    return jsonify(result)


@app.route('/api/cleaning/preview-table/<filename>')
def api_cleaning_preview_table(filename):
    """
    GET preview tabel df_current (10 baris pertama) sebagai HTML.
    Dipakai untuk refresh preview di Data Preview tab setelah cleaning.
    """
    sess, err = _get_or_init_session(filename)
    if err:
        return jsonify({'ok': False, 'error': err}), 404

    df   = sess.df_current
    html = df.to_html(
        classes='data-table display nowrap',
        index=False,
        border=0,
        table_id='preview-clean-table',
    )
    return jsonify({
        'ok'          : True,
        'html'        : html,
        'is_cleaned'  : sess.is_cleaned,
        'rows'        : len(df),
        'cols'        : len(df.columns),
        'missing'     : int(df.isna().sum().sum()),
    })


@app.route('/api/cleaning/columns/<filename>')
def api_cleaning_columns(filename):
    """
    GET daftar kolom df_current beserta info (dtype, missing, is_numeric).
    Dipakai untuk populate dropdown kolom di form cleaning.
    """
    sess, err = _get_or_init_session(filename)
    if err:
        return jsonify({'ok': False, 'error': err}), 404

    df   = sess.df_current
    cols = []
    for col in df.columns:
        s = df[col]
        cols.append({
            'name'       : col,
            'dtype'      : str(s.dtype),
            'missing'    : int(s.isna().sum()),
            'missing_pct': round(s.isna().mean() * 100, 1),
            'is_numeric' : bool(pd.api.types.is_numeric_dtype(s)),
            'is_text'    : bool(pd.api.types.is_object_dtype(s)),
            'unique'     : int(s.nunique()),
        })
    return jsonify({'ok': True, 'columns': cols, 'is_cleaned': sess.is_cleaned})


# ─────────────────────────────────────────────────────────────────────────────
# VISUALIZATION MASTER API
# ─────────────────────────────────────────────────────────────────────────────

def _load_analysis_df(filename):
    """
    Muat df untuk visualisasi.
    Prioritaskan df_current dari CleaningSession jika sudah cleaned,
    fallback ke clean_dataset(df_raw, {}) untuk backward-compat.
    """
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if not os.path.exists(filepath):
        return None, None, None, None

    # Cek session — pakai df yang sudah cleaned jika ada
    sess = get_session(filename)
    if sess is not None:
        df = sess.df_current
    else:
        df_raw = load_data(filepath)
        if df_raw is None:
            return None, None, None, None
        df, _ = clean_dataset(df_raw, {})

    num_cols, cat_cols = detect_data_types(df)
    dt_cols = detect_datetime_cols(df)
    return df, num_cols, cat_cols, dt_cols


@app.route('/api/viz-chart/<filename>')
def api_viz_chart(filename):
    """Generate satu grafik master berdasarkan kategori, tipe, dan kolom."""
    df, num_cols, cat_cols, _ = _load_analysis_df(filename)
    if df is None:
        return jsonify({'ok': False, 'placeholder': 'Dataset tidak ditemukan.'}), 404

    category   = request.args.get('category', 'numerical')
    chart_type = request.args.get('chart_type', '')
    col_x      = request.args.get('col_x') or None
    col_y      = request.args.get('col_y') or None
    col_z      = request.args.get('col_z') or None

    if not chart_type:
        types      = CATEGORY_CHARTS.get(category, [])
        chart_type = types[0] if types else ''

    result = generate_master_chart(
        df, num_cols, cat_cols, category, chart_type,
        col_x=col_x, col_y=col_y, col_z=col_z,
    )
    result['category']  = category
    result['available'] = category_available(category, num_cols, cat_cols)
    return jsonify(result)


# ─────────────────────────────────────────────────────────────────────────────
# AI CHAT ENDPOINT
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/api/ai-chat', methods=['POST'])
def ai_chat():
    """Proxy-free AI endpoint — pakai Anthropic API langsung dari server."""
    try:
        import anthropic
        data     = request.get_json()
        user_msg = data.get('message', '')
        context  = data.get('context', '')

        client = anthropic.Anthropic(api_key=os.environ.get('ANTHROPIC_API_KEY', ''))
        system_prompt = (
            "Kamu adalah asisten analisis data cerdas bernama DataBot untuk aplikasi "
            "DS Generator Kelompok 2. "
            "Jawab dengan bahasa yang jelas, ringkas, dan informatif. "
            f"Konteks dataset saat ini: {context}"
        )
        message = client.messages.create(
            model      = "claude-sonnet-4-20250514",
            max_tokens = 1024,
            system     = system_prompt,
            messages   = [{"role": "user", "content": user_msg}],
        )
        return jsonify({'reply': message.content[0].text})

    except ImportError:
        return jsonify({
            'reply': 'Library anthropic belum terpasang. Jalankan: pip install anthropic'
        }), 500
    except Exception as e:
        return jsonify({'reply': f'Error: {str(e)}'}), 500


# ─────────────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    app.run(debug=True)